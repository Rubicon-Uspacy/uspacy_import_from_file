#!/usr/bin/env python3
import argparse
import csv
import os
import sys
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional

import requests
from openpyxl import load_workbook


@dataclass
class FieldInfo:
    field_type: str
    list_title_to_value: Dict[str, str]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Update Uspacy entities from a CSV/XLSX file. The first row must contain field IDs "
            "as in Uspacy, and the first column is used to search entities."
        )
    )
    parser.add_argument(
        "--base-url",
        required=True,
        help="Base URL like https://{domain}.uspacy.ua",
    )
    parser.add_argument(
        "--entity",
        required=True,
        help="Entity name, e.g., companies or contacts",
    )
    parser.add_argument(
        "--file",
        required=True,
        help="Path to CSV/XLSX file (first row is field IDs)",
    )
    parser.add_argument(
        "--search-field",
        help="Field ID for lookup (defaults to the first column in the file)",
    )
    parser.add_argument(
        "--webhook-header",
        default="Authorization",
        help="Header name for webhook auth (default: Authorization)",
    )
    parser.add_argument(
        "--webhook-token",
        default=os.environ.get("USPACY_WEBHOOK_TOKEN"),
        help="Token for webhook auth (default: env USPACY_WEBHOOK_TOKEN)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Do not patch, only log what would be updated",
    )
    return parser.parse_args()


def load_rows(file_path: str) -> Iterable[List[str]]:
    if file_path.lower().endswith(".csv"):
        with open(file_path, newline="", encoding="utf-8") as handle:
            reader = csv.reader(handle)
            for row in reader:
                yield row
        return

    if file_path.lower().endswith(".xlsx"):
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            yield ["" if cell is None else str(cell).strip() for cell in row]
        return

    raise ValueError("Unsupported file type. Use .csv or .xlsx")


def request_session(webhook_header: str, webhook_token: Optional[str]) -> requests.Session:
    session = requests.Session()
    if webhook_token:
        session.headers[webhook_header] = webhook_token
    session.headers["Content-Type"] = "application/json"
    return session


def fetch_fields(
    session: requests.Session, api_base: str, entity: str
) -> Dict[str, FieldInfo]:
    url = f"{api_base}/crm/v1/entities/{entity}/fields"
    response = session.get(url, timeout=30)
    response.raise_for_status()
    data = response.json()
    fields: Dict[str, FieldInfo] = {}
    for field in data:
        field_id = field.get("id")
        field_type = field.get("type", "")
        title_map: Dict[str, str] = {}
        if field_type == "list":
            for value in field.get("values", []):
                title = str(value.get("title", "")).strip()
                val = str(value.get("value", "")).strip()
                if title:
                    title_map[title] = val
        if field_id:
            fields[str(field_id)] = FieldInfo(field_type=field_type, list_title_to_value=title_map)
    return fields


def build_update_payload(
    row: Dict[str, str],
    search_field: str,
    fields_info: Dict[str, FieldInfo],
) -> Dict[str, Any]:
    payload: Dict[str, Any] = {}
    for field_id, value in row.items():
        if field_id == search_field:
            continue
        if value == "":
            continue
        field_info = fields_info.get(field_id)
        if field_info and field_info.field_type == "list":
            mapped = field_info.list_title_to_value.get(value)
            if mapped is None:
                print(
                    f"[WARN] List field '{field_id}' value '{value}' not found in Uspacy list. Skipping.",
                    file=sys.stderr,
                )
                continue
            payload[field_id] = mapped
        else:
            payload[field_id] = value
    return payload


def search_entity(
    session: requests.Session,
    api_base: str,
    entity: str,
    search_field: str,
    search_value: str,
) -> List[Dict[str, Any]]:
    url = f"{api_base}/crm/v1/entities/{entity}/"
    params = {
        "boolean_operator": "AND",
        "page": 1,
        "list": 20,
        search_field: search_value,
    }
    response = session.get(url, params=params, timeout=30)
    response.raise_for_status()
    data = response.json()
    return data if isinstance(data, list) else data.get("data", [])


def patch_entity(
    session: requests.Session,
    api_base: str,
    entity: str,
    entity_id: str,
    payload: Dict[str, Any],
) -> None:
    url = f"{api_base}/crm/v1/entities/{entity}/{entity_id}"
    response = session.patch(url, json=payload, timeout=30)
    response.raise_for_status()


def main() -> int:
    args = parse_args()
    if not args.webhook_token:
        print("Webhook token is required. Use --webhook-token or set USPACY_WEBHOOK_TOKEN.")
        return 1

    rows_iter = load_rows(args.file)
    try:
        headers = next(rows_iter)
    except StopIteration:
        print("File is empty.")
        return 1

    if not headers or all(cell == "" for cell in headers):
        print("Header row is empty.")
        return 1

    search_field = args.search_field or headers[0]
    if search_field not in headers:
        print(f"Search field '{search_field}' not found in header row.")
        return 1

    base_url = args.base_url.rstrip("/")
    api_base = f"{base_url}/company/v1/incoming_webhooks/run/{args.webhook_token}"
    session = request_session(args.webhook_header, args.webhook_token)
    fields_info = fetch_fields(session, api_base, args.entity)

    for index, row_values in enumerate(rows_iter, start=2):
        if not any(cell != "" for cell in row_values):
            continue
        row = {headers[i]: row_values[i] if i < len(row_values) else "" for i in range(len(headers))}
        search_value = row.get(search_field, "")
        if search_value == "":
            print(f"[WARN] Row {index}: empty search value, skipping.")
            continue

        matches = search_entity(session, api_base, args.entity, search_field, search_value)
        if not matches:
            print(f"[WARN] Row {index}: no match for {search_field}={search_value}")
            continue

        if len(matches) > 1:
            print(
                f"[WARN] Row {index}: multiple matches for {search_field}={search_value}, using first.",
                file=sys.stderr,
            )
        entity_id = str(matches[0].get("id"))
        if entity_id in ("", "None"):
            print(f"[WARN] Row {index}: missing entity id for {search_field}={search_value}")
            continue

        payload = build_update_payload(row, search_field, fields_info)
        if not payload:
            print(f"[INFO] Row {index}: nothing to update for {search_field}={search_value}")
            continue

        if args.dry_run:
            print(f"[DRY-RUN] Row {index}: PATCH {entity_id} -> {payload}")
            continue

        patch_entity(session, api_base, args.entity, entity_id, payload)
        print(f"[OK] Row {index}: updated {entity_id}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
